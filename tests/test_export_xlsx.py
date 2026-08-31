#! python3  # noqa E265

"""
Usage from the repo root folder:

.. code-block:: bash
    # for whole tests
    python -m unittest tests.test_export_xlsx
    # for specific test
    python -m unittest tests.test_export_xlsx.TestFormatBbox.test_valid_tuple_is_comma_joined
"""

# standard library
import tempfile
import unittest
from functools import lru_cache
from pathlib import Path

# 3rd party
from openpyxl import load_workbook

# project
from dicogis.export.to_xlsx import MetadatasetSerializerXlsx
from dicogis.models.feature_attributes import AttributeField
from dicogis.models.metadataset import (
    MetaDatabaseFlat,
    MetaDatabaseTable,
    MetaDataset,
    MetaRasterDataset,
    MetaVectorDataset,
)
from dicogis.utils.texts import TextsManager

# ############################################################################
# ########## Globals #############
# ################################


@lru_cache
def _en_texts() -> dict:
    """Real EN localized strings, loaded once and shared across tests."""
    return TextsManager(locale_folder=Path("locale")).load_texts(language_code="EN")


def _make_serializer(**overrides) -> MetadatasetSerializerXlsx:
    kwargs = dict(localized_strings=_en_texts())
    kwargs.update(overrides)
    return MetadatasetSerializerXlsx(**kwargs)


# ############################################################################
# ########## Classes #############
# ################################


class TestFormatBbox(unittest.TestCase):
    """Test MetadatasetSerializerXlsx.format_bbox()."""

    def test_valid_tuple_is_comma_joined(self):
        serializer = _make_serializer()
        self.assertEqual(
            serializer.format_bbox(bbox=(1.0, 2.0, 3, 4)), "1.0, 2.0, 3, 4"
        )

    def test_none_returns_empty_string(self):
        serializer = _make_serializer()
        self.assertEqual(serializer.format_bbox(bbox=None), "")

    def test_non_tuple_returns_empty_string(self):
        serializer = _make_serializer()
        self.assertEqual(serializer.format_bbox(bbox="not-a-bbox"), "")

    def test_tuple_with_non_numeric_member_returns_empty_string(self):
        serializer = _make_serializer()
        self.assertEqual(serializer.format_bbox(bbox=(1.0, "oops", 3, 4)), "")


class TestFormatAsHyperlink(unittest.TestCase):
    """Test MetadatasetSerializerXlsx.format_as_hyperlink()."""

    def test_str_target(self):
        serializer = _make_serializer()
        self.assertEqual(
            serializer.format_as_hyperlink(target="C:/data", label="browse"),
            '=HYPERLINK("C:/data", "browse")',
        )

    def test_path_target_is_resolved(self):
        serializer = _make_serializer()
        result = serializer.format_as_hyperlink(target=Path("data"), label="browse")
        self.assertIn(str(Path("data").resolve()), result)
        self.assertTrue(result.startswith("=HYPERLINK("))


class TestFormatSize(unittest.TestCase):
    """Test MetadatasetSerializerXlsx.format_size()."""

    def test_prettify_enabled(self):
        serializer = _make_serializer(opt_size_prettify=True)
        self.assertEqual(serializer.format_size(in_size_in_octets=1024), "1.0 Ko")

    def test_prettify_disabled_returns_raw_value(self):
        serializer = _make_serializer(opt_size_prettify=False)
        self.assertEqual(serializer.format_size(in_size_in_octets=1024), 1024)


class TestFormatFeatureAttributes(unittest.TestCase):
    """Test MetadatasetSerializerXlsx.format_feature_attributes()."""

    def test_non_list_feature_attributes_returns_empty_string(self):
        serializer = _make_serializer()
        metadataset = MetaVectorDataset(feature_attributes=None)
        self.assertEqual(serializer.format_feature_attributes(metadataset), "")

    def test_known_types_are_translated(self):
        serializer = _make_serializer()
        metadataset = MetaVectorDataset(
            feature_attributes=[
                AttributeField(name="id", data_type="Integer", length=10, precision=0),
                AttributeField(name="area", data_type="Real", length=10, precision=2),
                AttributeField(name="label", data_type="String", length=50),
                AttributeField(name="created", data_type="Date"),
                AttributeField(name="blob", data_type="Binary"),
            ]
        )
        result = serializer.format_feature_attributes(metadataset)
        for expected_field_name in ("id", "area", "label", "created", "blob"):
            self.assertIn(expected_field_name, result)
        # translated type labels show up instead of the raw GDAL type names
        self.assertNotIn("Integer", result)
        self.assertNotIn("Real (", result)

    def test_unknown_type_falls_back_to_raw_value(self):
        serializer = _make_serializer()
        metadataset = MetaVectorDataset(
            feature_attributes=[
                AttributeField(name="mystery", data_type="WeirdCustomType")
            ]
        )
        result = serializer.format_feature_attributes(metadataset)
        self.assertIn("mystery", result)
        self.assertIn("WeirdCustomType", result)


class TestIsStyleRegistered(unittest.TestCase):
    """Test the (dead) MetadatasetSerializerXlsx.is_style_registered()."""

    def test_registered_style_is_found(self):
        serializer = _make_serializer()

        self.assertTrue(serializer.is_style_registered(style_name="date"))
        self.assertTrue(serializer.is_style_registered(style_name="wrap"))

    def test_unregistered_style_is_not_found(self):
        serializer = _make_serializer()

        self.assertFalse(serializer.is_style_registered(style_name="not-a-style"))


class TestPreSerializing(unittest.TestCase):
    """Test MetadatasetSerializerXlsx.pre_serializing()."""

    def test_creates_vector_sheet_with_translated_headers(self):
        serializer = _make_serializer()
        serializer.pre_serializing(has_vector=True)

        self.assertIn("Vectors", serializer.workbook.sheetnames)
        header_row = [cell.value for cell in serializer.sheet_vector_files[1]]
        self.assertEqual(header_row[0], "Filename")
        self.assertEqual(serializer.row_index_vector_files, 1)

    def test_creates_raster_and_filedb_sheets(self):
        serializer = _make_serializer()
        serializer.pre_serializing(has_raster=True, has_filedb=True)

        self.assertIn("Rasters", serializer.workbook.sheetnames)
        self.assertIn("File databases", serializer.workbook.sheetnames)

    def test_creates_sgbd_sheet_with_translated_headers(self):
        serializer = _make_serializer()
        serializer.pre_serializing(has_sgbd=True)

        self.assertIn("PostGIS", serializer.workbook.sheetnames)
        header_row = [cell.value for cell in serializer.sheet_server_geodatabases[1]]
        self.assertEqual(header_row[-1], "GDAL errors")

    def test_calling_twice_does_not_duplicate_sheet(self):
        serializer = _make_serializer()
        serializer.pre_serializing(has_vector=True)
        serializer.pre_serializing(has_vector=True)

        self.assertEqual(serializer.workbook.sheetnames.count("Vectors"), 1)

    def test_creates_cad_sheet_with_translated_headers(self):
        serializer = _make_serializer()
        serializer.pre_serializing(has_cad=True)

        self.assertIn("CAD", serializer.workbook.sheetnames)
        header_row = [cell.value for cell in serializer.sheet_cad_files[1]]
        self.assertEqual(header_row[0], "Filename")
        self.assertEqual(serializer.row_index_cad_files, 1)


class TestGetSheetAndIncrementedRowIndexFromType(unittest.TestCase):
    """Test MetadatasetSerializerXlsx.get_sheet_and_incremented_row_index_from_type()."""

    def setUp(self):
        self.serializer = _make_serializer()
        self.serializer.pre_serializing(
            has_vector=True, has_raster=True, has_filedb=True, has_sgbd=True
        )

    def test_routes_vector_dataset_and_increments_row(self):
        metadataset = MetaVectorDataset(dataset_type="flat_vector")
        sheet, row_index = (
            self.serializer.get_sheet_and_incremented_row_index_from_type(metadataset)
        )
        self.assertIs(sheet, self.serializer.sheet_vector_files)
        self.assertEqual(row_index, 2)

    def test_routes_raster_dataset_regardless_of_dataset_type(self):
        metadataset = MetaRasterDataset()
        sheet, row_index = (
            self.serializer.get_sheet_and_incremented_row_index_from_type(metadataset)
        )
        self.assertIs(sheet, self.serializer.sheet_raster_files)
        self.assertEqual(row_index, 2)

    def test_routes_database_table_to_server_geodatabases(self):
        metadataset = MetaDatabaseTable(dataset_type="sgbd_postgis")
        sheet, row_index = (
            self.serializer.get_sheet_and_incremented_row_index_from_type(metadataset)
        )
        self.assertIs(sheet, self.serializer.sheet_server_geodatabases)
        self.assertEqual(row_index, 2)

    def test_routes_database_flat_to_flat_geodatabases(self):
        metadataset = MetaDatabaseFlat(dataset_type="flat_database")
        sheet, row_index = (
            self.serializer.get_sheet_and_incremented_row_index_from_type(metadataset)
        )
        self.assertIs(sheet, self.serializer.sheet_flat_geodatabases)
        self.assertEqual(row_index, 2)

    def test_repeated_calls_increment_row_index(self):
        metadataset = MetaVectorDataset(dataset_type="flat_vector")
        self.serializer.get_sheet_and_incremented_row_index_from_type(metadataset)
        _, row_index = self.serializer.get_sheet_and_incremented_row_index_from_type(
            metadataset
        )
        self.assertEqual(row_index, 3)

    def test_routes_cad_dataset_and_increments_row(self):
        metadataset = MetaVectorDataset(dataset_type="flat_cad")
        self.serializer.pre_serializing(has_cad=True)

        sheet, row_index = (
            self.serializer.get_sheet_and_incremented_row_index_from_type(metadataset)
        )

        self.assertIs(sheet, self.serializer.sheet_cad_files)
        self.assertEqual(row_index, 2)

    def test_unmatched_type_raises_clear_valueerror(self):
        """A plain ``MetaDataset`` (or a ``MetaVectorDataset`` whose
        ``dataset_type`` isn't a routed one) matches none of the
        isinstance/dataset_type branches, so a clear ``ValueError`` is
        raised instead of an implicit ``None`` that would fail confusingly
        on unpacking in the caller."""
        metadataset = MetaDataset()

        with self.assertRaises(ValueError):
            self.serializer.get_sheet_and_incremented_row_index_from_type(metadataset)


class TestStoreError(unittest.TestCase):
    """Test MetadatasetSerializerXlsx.store_error()."""

    def test_stores_name_hyperlink_and_message_with_warning_style(self):
        serializer = _make_serializer()
        serializer.pre_serializing(has_vector=True)
        worksheet = serializer.sheet_vector_files
        metadataset = MetaDataset(
            name="broken.shp",
            path=Path("broken.shp"),
            processing_error_type="err_unknown_format",
            processing_error_msg="unrecognized driver",
        )

        serializer.store_error(
            metadataset=metadataset, worksheet=worksheet, row_index=2
        )

        self.assertEqual(worksheet["A2"].value, "broken.shp")
        self.assertEqual(worksheet["A2"].style, "Warning Text")
        self.assertIn("HYPERLINK", worksheet["B2"].value)
        self.assertEqual(worksheet["C2"].value, "unrecognized driver")
        # gdal info goes into the sheet's last column (Q, "gdal_warn" here)
        self.assertIn("unrecognized driver", worksheet["Q2"].value)

    def test_error_lands_in_the_sheets_own_last_column(self):
        """store_error() finds the gdal-warning column dynamically (the
        sheet's own last column), rather than assuming a hardcoded "Q" --
        which only happened to be correct for the vector sheet. On the
        raster sheet, that's V (li_cols_raster's real "gdal_warn" column),
        not Q ("format")."""
        serializer = _make_serializer()
        serializer.pre_serializing(has_raster=True)
        metadataset = MetaRasterDataset(
            name="broken.tif",
            path=Path("broken.tif"),
            processing_error_type="err_unknown_format",
            processing_error_msg="unrecognized driver",
        )

        serializer.store_error(
            metadataset=metadataset,
            worksheet=serializer.sheet_raster_files,
            row_index=2,
        )

        raster_headers = [cell.value for cell in serializer.sheet_raster_files[1]]
        self.assertEqual(raster_headers[16], "Format")  # column Q, 0-indexed 16
        self.assertEqual(raster_headers[21], "GDAL warnings")  # column V
        self.assertIsNone(serializer.sheet_raster_files["Q2"].value)
        self.assertIn("unrecognized driver", serializer.sheet_raster_files["V2"].value)


class TestSerializeMetadatasetVector(unittest.TestCase):
    """Test MetadatasetSerializerXlsx.serialize_metadaset() for vector datasets."""

    def setUp(self):
        self.serializer = _make_serializer()
        self.serializer.pre_serializing(has_vector=True)

    def _vector_metadataset(self, **overrides) -> MetaVectorDataset:
        kwargs = dict(
            name="parcels.shp",
            path=Path("data/parcels.shp"),
            parent_folder_name="data",
            dataset_type="flat_vector",
            feature_attributes=[AttributeField(name="id", data_type="Integer")],
            features_objects_count=42,
            geometry_type="Polygon",
            crs_name="RGF93",
            crs_type="projected",
            crs_registry_code="2154",
            bbox=(0.0, 0.0, 1.0, 1.0),
            format_gdal_long_name="ESRI Shapefile",
            files_dependencies=[],
            storage_size=2048,
        )
        kwargs.update(overrides)
        return MetaVectorDataset(**kwargs)

    def test_happy_path_writes_expected_columns(self):
        metadataset = self._vector_metadataset()

        self.serializer.serialize_metadaset(metadataset=metadataset)

        sheet = self.serializer.sheet_vector_files
        self.assertEqual(sheet["A2"].value, "parcels.shp")
        self.assertEqual(sheet["C2"].value, "data")
        self.assertEqual(sheet["D2"].value, 1)  # count_feature_attributes
        self.assertEqual(sheet["E2"].value, 42)  # features_objects_count
        self.assertEqual(sheet["F2"].value, "Polygon")  # geometry_type
        self.assertEqual(sheet["G2"].value, "RGF93")  # crs_name
        self.assertEqual(sheet["I2"].value, "EPSG:2154")  # codepsg
        self.assertEqual(sheet["M2"].value, "ESRI Shapefile")  # format
        self.assertEqual(sheet["O2"].value, "2.0 Ko")  # tot_size, prettified

    def test_raw_path_option_skips_hyperlink(self):
        serializer = _make_serializer(opt_raw_path=True)
        serializer.pre_serializing(has_vector=True)
        metadataset = self._vector_metadataset()

        serializer.serialize_metadaset(metadataset=metadataset)

        self.assertEqual(
            serializer.sheet_vector_files["B2"].value,
            str(Path("data/parcels.shp").resolve()),
        )

    def test_dependencies_are_joined_as_strings_in_correct_column(self):
        metadataset = self._vector_metadataset(
            files_dependencies=[Path("data/parcels.dbf"), Path("data/parcels.shx")]
        )

        self.serializer.serialize_metadaset(metadataset=metadataset)

        dependencies_cell = self.serializer.sheet_vector_files["N2"].value
        self.assertIn("parcels.dbf", dependencies_cell)
        self.assertIn("parcels.shx", dependencies_cell)

    def test_processing_error_stores_message_and_warning_style_after_normal_write(self):
        """A processing failure is flagged *after* the normal vector-row
        write, so the warning highlighting and error message survive
        instead of being immediately clobbered by it (previously, C ended
        up showing the folder name -- not the error -- while still
        highlighted as a warning)."""
        metadataset = self._vector_metadataset(
            processing_succeeded=False,
            processing_error_type="err_unknown",
            processing_error_msg="could not open dataset",
        )

        self.serializer.serialize_metadaset(metadataset=metadataset)

        sheet = self.serializer.sheet_vector_files
        self.assertEqual(sheet["A2"].style, "Warning Text")
        self.assertEqual(sheet["B2"].style, "Warning Text")
        self.assertEqual(sheet["C2"].value, "could not open dataset")
        self.assertEqual(sheet["C2"].style, "Warning Text")
        self.assertIn("could not open dataset", sheet["Q2"].value)
        # normal vector fields were still written
        self.assertEqual(sheet["F2"].value, "Polygon")

    def test_missing_storage_size_is_left_blank_instead_of_crashing(self):
        """storage_size=None (the dataclass default) degrades gracefully to
        an empty size cell instead of crashing serialization."""
        metadataset = self._vector_metadataset(storage_size=None)

        self.serializer.serialize_metadaset(metadataset=metadataset)

        self.assertEqual(self.serializer.sheet_vector_files["O2"].value, "")


class TestSerializeMetadatasetCad(unittest.TestCase):
    """Test MetadatasetSerializerXlsx.serialize_metadaset() for CAD/DAO datasets."""

    def test_cad_dataset_is_routed_to_the_cad_sheet(self):
        serializer = _make_serializer()
        serializer.pre_serializing(has_cad=True)
        metadataset = MetaVectorDataset(
            name="plan.dxf",
            path=Path("data/plan.dxf"),
            parent_folder_name="data",
            dataset_type="flat_cad",
            feature_attributes=[AttributeField(name="id", data_type="Integer")],
            geometry_type="LineString",
            files_dependencies=[],
            storage_size=512,
        )

        serializer.serialize_metadaset(metadataset=metadataset)

        sheet = serializer.sheet_cad_files
        self.assertEqual(sheet["A2"].value, "plan.dxf")
        self.assertEqual(sheet["F2"].value, "LineString")
        self.assertEqual(sheet["O2"].value, "512.0 octets")


class TestSerializeMetadatasetRaster(unittest.TestCase):
    """Test MetadatasetSerializerXlsx.serialize_metadaset() for raster datasets."""

    def setUp(self):
        self.serializer = _make_serializer()
        self.serializer.pre_serializing(has_raster=True)

    def _raster_metadataset(self, **overrides) -> MetaRasterDataset:
        kwargs = dict(
            name="ortho.tif",
            path=Path("data/ortho.tif"),
            parent_folder_name="data",
            dataset_type="flat_raster",
            rows_count=100,
            columns_count=200,
            pixel_width=0.5,
            pixel_height=0.5,
            crs_name="RGF93",
            format_gdal_long_name="GeoTIFF",
            files_dependencies=[],
            storage_size=4096,
        )
        kwargs.update(overrides)
        return MetaRasterDataset(**kwargs)

    def test_happy_path_writes_expected_columns(self):
        metadataset = self._raster_metadataset()

        self.serializer.serialize_metadaset(metadataset=metadataset)

        sheet = self.serializer.sheet_raster_files
        self.assertEqual(sheet["D2"].value, 100)  # rows_count
        self.assertEqual(sheet["E2"].value, 200)  # columns_count
        self.assertEqual(sheet["J2"].value, "RGF93")  # crs_name
        self.assertEqual(sheet["Q2"].value, "GeoTIFF")  # format

    def test_dependencies_and_size_land_in_correct_columns(self):
        """T (li_depends) holds the dependency text and U (tot_size) holds
        the formatted size -- previously shifted one column each, with the
        size ending up mislabeled under the gdal_warn column (V)."""
        metadataset = self._raster_metadataset(files_dependencies=[], storage_size=1024)

        self.serializer.serialize_metadaset(metadataset=metadataset)

        sheet = self.serializer.sheet_raster_files
        self.assertEqual(sheet["T2"].value, "")
        self.assertEqual(sheet["U2"].value, "1.0 Ko")
        self.assertIsNone(sheet["V2"].value)

    def test_dependencies_are_joined_as_strings(self):
        """Dependencies are stringified (str(f.resolve())) like the vector
        writer does, instead of joining raw Path objects (which used to
        raise TypeError as soon as there was at least one dependency)."""
        metadataset = self._raster_metadataset(
            files_dependencies=[Path("data/ortho.tfw")]
        )

        self.serializer.serialize_metadaset(metadataset=metadataset)

        self.assertIn("ortho.tfw", self.serializer.sheet_raster_files["T2"].value)


class TestStoreMdFlatGeodatabases(unittest.TestCase):
    """Test MetadatasetSerializerXlsx.store_md_flat_geodatabases()."""

    def setUp(self):
        self.serializer = _make_serializer()
        self.serializer.pre_serializing(has_filedb=True)

    def test_dataset_without_layers_logs_and_returns_early(self):
        metadataset = MetaDatabaseFlat(
            name="empty.gdb",
            path=Path("empty.gdb"),
            parent_folder_name="data",
            layers=None,
            storage_size=10,
        )

        # mirror what get_sheet_and_incremented_row_index_from_type() would
        # have done: keep the instance counter in sync with row_index
        self.serializer.row_index_flat_geodatabases = 2

        # should not raise despite there being no layers to iterate
        self.serializer.store_md_flat_geodatabases(
            metadataset=metadataset,
            worksheet=self.serializer.sheet_flat_geodatabases,
            row_index=2,
        )

        self.assertEqual(self.serializer.row_index_flat_geodatabases, 2)

    def test_dataset_with_layers_writes_one_row_per_layer(self):
        layers = [
            MetaVectorDataset(
                name="roads",
                feature_attributes=[AttributeField(name="id", data_type="Integer")],
                features_objects_count=5,
                geometry_type="LineString",
            ),
            MetaVectorDataset(
                name="buildings",
                feature_attributes=[AttributeField(name="id", data_type="Integer")],
                features_objects_count=9,
                geometry_type="Polygon",
            ),
        ]
        metadataset = MetaDatabaseFlat(
            name="city.gdb",
            path=Path("city.gdb"),
            parent_folder_name="data",
            layers=layers,
            storage_size=10,
        )

        # mirror what get_sheet_and_incremented_row_index_from_type() would
        # have done: keep the instance counter in sync with row_index
        self.serializer.row_index_flat_geodatabases = 2

        self.serializer.store_md_flat_geodatabases(
            metadataset=metadataset,
            worksheet=self.serializer.sheet_flat_geodatabases,
            row_index=2,
        )

        sheet = self.serializer.sheet_flat_geodatabases
        self.assertEqual(sheet["H2"].value, 2)  # count_layers, on the summary row
        self.assertEqual(sheet["H3"].value, "roads")
        self.assertEqual(sheet["K3"].value, "LineString")
        self.assertEqual(sheet["H4"].value, "buildings")
        self.assertEqual(sheet["K4"].value, "Polygon")


class TestStoreMdGeodatabasesServer(unittest.TestCase):
    """Test MetadatasetSerializerXlsx.store_md_geodatabases_server()."""

    def test_writes_expected_columns(self):
        serializer = _make_serializer()
        serializer.pre_serializing(has_sgbd=True)
        metadataset = MetaDatabaseTable(
            name="public.roads",
            description="Road network reference table",
            dataset_type="sgbd_postgis",
            schema_name="public",
            feature_attributes=[AttributeField(name="id", data_type="Integer")],
            features_objects_count=7,
            geometry_type="LineString",
            crs_name="RGF93",
            crs_registry_code="2154",
        )

        serializer.serialize_metadaset(metadataset=metadataset)

        sheet = serializer.sheet_server_geodatabases
        self.assertEqual(sheet["C2"].value, "public")
        self.assertEqual(sheet["D2"].value, "Road network reference table")
        self.assertEqual(sheet["E2"].value, 1)
        self.assertEqual(sheet["F2"].value, 7)
        self.assertEqual(sheet["G2"].value, "LineString")
        self.assertEqual(sheet["H2"].value, "RGF93")


class TestPostSerializingAndTunningWorkbook(unittest.TestCase):
    """End-to-end test: build a small workbook and save it to disk."""

    def test_saved_workbook_has_frozen_panes_and_filters(self):
        serializer = _make_serializer()
        serializer.pre_serializing(has_vector=True)
        metadataset = MetaVectorDataset(
            name="parcels.shp",
            path=Path("data/parcels.shp"),
            parent_folder_name="data",
            dataset_type="flat_vector",
            files_dependencies=[],
            storage_size=1024,
        )
        serializer.serialize_metadaset(metadataset=metadataset)

        with tempfile.TemporaryDirectory(
            prefix="DicoGIS_test_export_xlsx_", ignore_cleanup_errors=True
        ) as tmpdirname:
            output_path = Path(tmpdirname) / "output.xlsx"
            serializer.output_path = output_path
            serializer.post_serializing()

            self.assertTrue(output_path.is_file())

            reloaded = load_workbook(filename=output_path)
            self.assertIn("Vectors", reloaded.sheetnames)
            sheet = reloaded["Vectors"]
            self.assertEqual(sheet.freeze_panes, "B2")
            self.assertEqual(sheet.auto_filter.ref, f"A1:Q{sheet.max_row}")
            self.assertEqual(sheet["A1"].style, "Headline 2")


# ############################################################################
# ####### Stand-alone run ########
# ################################
if __name__ == "__main__":
    unittest.main()

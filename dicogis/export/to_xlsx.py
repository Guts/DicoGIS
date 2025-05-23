#! python3  # noqa: E265

# ##############################################################################
# ########## Libraries #############
# ##################################

# Standard library
import logging
from collections.abc import Iterable
from functools import lru_cache
from pathlib import Path
from typing import Union

# 3rd party library
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# project
from dicogis.export.base_serializer import MetadatasetSerializerBase
from dicogis.models.metadataset import (
    MetaDatabaseFlat,
    MetaDatabaseTable,
    MetaDataset,
    MetaRasterDataset,
    MetaVectorDataset,
)
from dicogis.utils.formatters import convert_octets

# ##############################################################################
# ############ Globals ############
# #################################

# LOG
logger = logging.getLogger(__name__)


# ##############################################################################
# ########## Classes ###############
# ##################################


class MetadatasetSerializerXlsx(MetadatasetSerializerBase):
    """Export into a XLSX worksheet."""

    li_cols_vector = [
        "nomfic",
        "path",
        "theme",
        "num_attrib",
        "num_objets",
        "geometrie",
        "srs",
        "srs_type",
        "codepsg",
        "emprise",
        "date_crea",
        "date_actu",
        "format",
        "li_depends",
        "tot_size",
        "li_chps",
        "gdal_warn",
    ]

    li_cols_raster = [
        "nomfic",
        "path",
        "theme",
        "size_Y",
        "size_X",
        "pixel_w",
        "pixel_h",
        "origin_x",
        "origin_y",  # I
        "srs",  # J
        "srs_type",
        "codepsg",
        "emprise",
        "date_crea",
        "date_actu",
        "num_bands",
        "format",
        "compression",
        "coloref",
        "li_depends",
        "tot_size",
        "gdal_warn",
    ]

    li_cols_filedb = [
        "nomfic",
        "path",
        "theme",
        "tot_size",
        "date_crea",
        "date_actu",
        "format",
        "feats_class",
        "num_attrib",
        "num_objets",
        "geometrie",
        "srs",
        "srs_type",
        "codepsg",
        "emprise",
        "li_chps",
    ]

    li_cols_mapdocs = [
        "nomfic",
        "path",
        "theme",
        "custom_title",
        "creator_prod",
        "keywords",
        "subject",
        "res_image",
        "tot_size",
        "date_crea",
        "date_actu",
        "origin_x",
        "origin_y",
        "srs",
        "srs_type",
        "codepsg",
        "sub_layers",
        "num_attrib",
        "num_objets",
        "li_chps",
    ]

    li_cols_sgbd = [
        "nomfic",
        "conn_chain",
        "schema",
        "num_attrib",
        "num_objets",
        "geometrie",
        "srs",
        "srs_type",
        "codepsg",
        "emprise",
        "format",
        "li_chps",
        "gdal_err",
    ]

    def __init__(
        self,
        # inherited
        localized_strings: dict | None = None,
        output_path: Path | None = None,
        opt_raw_path: bool = False,
        opt_size_prettify: bool = True,
    ) -> None:
        """Store metadata into Excel worksheets.

        Args:
            texts (dict, optional): dictionary of translated texts. Defaults to {}.
            opt_size_prettify (bool, optional): option to prettify size or not. Defaults to False.
        """
        self.workbook = Workbook(iso_dates=True)
        # styles
        s_date = NamedStyle(name="date")
        s_date.number_format = "dd/mm/yyyy"
        s_wrap = NamedStyle(name="wrap")
        s_wrap.alignment = Alignment(wrap_text=True)
        self.workbook.add_named_style(s_date)
        self.workbook.add_named_style(s_wrap)

        # deleting the default worksheet
        ws = self.workbook.active
        self.workbook.remove(worksheet=ws)

        # initiate with parent
        super().__init__(
            localized_strings=localized_strings,
            output_path=output_path,
            opt_raw_path=opt_raw_path,
            opt_size_prettify=opt_size_prettify,
        )

    # ------------ Setting workbook ---------------------

    def pre_serializing(
        self,
        has_vector: bool = False,
        has_raster: bool = False,
        has_filedb: bool = False,
        has_mapdocs: bool = False,
        has_cad: bool = False,
        has_sgbd: bool = False,
    ):
        """Prepare workbooks.

        Args:
            has_vector (bool, optional): _description_. Defaults to False.
            has_raster (bool, optional): _description_. Defaults to False.
            has_filedb (bool, optional): _description_. Defaults to False.
            has_mapdocs (bool, optional): _description_. Defaults to False.
            has_cad (bool, optional): _description_. Defaults to False.
            has_sgbd (bool, optional): _description_. Defaults to False.
        """

        # SHEETS & HEADERS
        if (
            has_vector
            and self.localized_strings.get("sheet_vectors")
            not in self.workbook.sheetnames
        ):
            self.sheet_vector_files = self.workbook.create_sheet(
                title=self.localized_strings.get("sheet_vectors")
            )
            # headers
            self.sheet_vector_files.append(
                [self.localized_strings.get(i) for i in self.li_cols_vector]
            )

            # initialize line counter
            self.row_index_vector_files = 1

        if (
            has_raster
            and self.localized_strings.get("sheet_rasters")
            not in self.workbook.sheetnames
        ):
            self.sheet_raster_files = self.workbook.create_sheet(
                title=self.localized_strings.get("sheet_rasters")
            )
            # headers
            self.sheet_raster_files.append(
                [self.localized_strings.get(i) for i in self.li_cols_raster]
            )

            # initialize line counter
            self.row_index_raster_files = 1

        if (
            has_filedb
            and self.localized_strings.get("sheet_filedb")
            not in self.workbook.sheetnames
        ):
            self.sheet_flat_geodatabases = self.workbook.create_sheet(
                title=self.localized_strings.get("sheet_filedb")
            )
            # headers
            self.sheet_flat_geodatabases.append(
                [self.localized_strings.get(i) for i in self.li_cols_filedb]
            )

            # initialize line counter
            self.row_index_flat_geodatabases = 1

        if (
            has_mapdocs
            and self.localized_strings.get("sheet_maplans")
            not in self.workbook.sheetnames
        ):
            self.sheet_map_workspaces = self.workbook.create_sheet(
                title=self.localized_strings.get("sheet_maplans")
            )
            # headers
            self.sheet_map_workspaces.append(
                [self.localized_strings.get(i) for i in self.li_cols_mapdocs]
            )

            # initialize line counter
            self.row_index_map_worskpaces = 1

        if (
            has_cad
            and self.localized_strings.get("sheet_cdao") not in self.workbook.sheetnames
        ):
            self.sheet_cad_files = self.workbook.create_sheet(
                title=self.localized_strings.get("sheet_cdao")
            )
            # headers
            self.sheet_cad_files.append(
                [self.localized_strings.get(i) for i in self.li_cols_caodao]
            )

            # initialize line counter
            self.row_index_cad_files = 1

        if has_sgbd and "PostGIS" not in self.workbook.sheetnames:
            self.sheet_server_geodatabases = self.workbook.create_sheet(title="PostGIS")
            # headers
            self.sheet_server_geodatabases.append(
                [self.localized_strings.get(i) for i in self.li_cols_sgbd]
            )
            # initialize line counter
            self.row_index_server_geodatabases = 1

    def post_serializing(self):
        """Run post serialization steps."""
        self.tunning_workbook()
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(filename=self.output_path)
        logger.info(f"Workbook saved under {self.output_path}")

    def tunning_workbook(self):
        """Clean up and tunning worksheet."""
        for sheet in self.workbook.worksheets:
            # Freezing panes
            c_freezed = sheet["B2"]
            sheet.freeze_panes = c_freezed

            # Print properties
            sheet.print_options.horizontalCentered = True
            sheet.print_options.verticalCentered = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

            # Others properties
            wsprops = sheet.sheet_properties
            wsprops.filterMode = True

            # headers
            for cell in sheet[1]:
                if cell.value:
                    cell.style = "Headline 2"

            # enable filters
            sheet.auto_filter.ref = "A1:{}{}".format(
                get_column_letter(sheet.max_column), sheet.max_row
            )
            # columns width
            sheet.column_dimensions["A"].bestFit = True
            sheet.column_dimensions["A"].auto_size = True
            sheet.column_dimensions["B"].auto_size = True

            # dims = {}
            # for row in sheet.rows:
            #     for cell in row:
            #         if cell.value:
            #             val = cell.value
            #             dims[cell.column] = max((dims.get(cell.column, 0), len(val)))
            # for col, value in dims.items():
            #     sheet.column_dimensions[col].width = value

    @lru_cache
    def is_style_registered(self, style_name: str) -> bool:
        """Function to check if a named style already exists/

        Args:
            style_name: name of style to check

        Returns:
            True if a style with the same name already exist in the worksheet
        """
        return any(style.name == style_name for style in self.workbook.named_styles)

    @lru_cache(maxsize=128, typed=True)
    def format_as_hyperlink(self, target: Union[str, Path], label: str) -> str:
        """Format a cell hyperlink with a target and a label.

        Args:
            target (Union[str, Path]): link destination
            label (str): display text

        Returns:
            str: cell text formatted with hyperlink
        """
        if isinstance(target, Path):
            target = str(target.resolve())

        return f'=HYPERLINK("{target}", "{label}")'

    @lru_cache(maxsize=128, typed=True)
    def format_bbox(self, bbox: tuple[float, float, float, float] | None) -> str:
        """Format bounding-box tuple to fit Excel constraints.

        Args:
            bbox: input bounding-box tuple

        Returns:
            formatted bbox as Excel-compliant string
        """
        if isinstance(bbox, tuple) and all(
            [isinstance(coord, (float, int)) for coord in bbox]
        ):
            return ", ".join(str(coord) for coord in bbox)
        else:
            return ""

    def format_feature_attributes(self, metadataset: MetaVectorDataset) -> str:
        """Format vector feature attributes in an unique string.

        Args:
            metadataset (MetaDataset): metadataset

        Returns:
            str: concatenated string with feature attributes informations
        """
        out_attributes_str = ""

        if not isinstance(metadataset.feature_attributes, (list, tuple)):
            return out_attributes_str

        for feature_attribute in metadataset.feature_attributes:
            # field type
            if "integer" in feature_attribute.data_type.lower():
                translated_feature_attribute_type = self.localized_strings.get("entier")
            elif feature_attribute.data_type.lower() == "real":
                translated_feature_attribute_type = self.localized_strings.get("reel")
            elif feature_attribute.data_type.lower() == "string":
                translated_feature_attribute_type = self.localized_strings.get("string")
            elif feature_attribute.data_type.lower() in ("date", "datetime"):
                translated_feature_attribute_type = self.localized_strings.get("date")
            elif feature_attribute.data_type.lower() == "binary":
                translated_feature_attribute_type = self.localized_strings.get("binary")
            else:
                translated_feature_attribute_type = feature_attribute.data_type
                logger.warning(
                    f"Layer: {metadataset.name} - {feature_attribute.name} has an "
                    f"unstranslated type: {feature_attribute.data_type}"
                )

            # concatenation of field informations
            out_attributes_str = "{} {} ({}{}{}{}{}); ".format(
                out_attributes_str,
                feature_attribute.name,
                translated_feature_attribute_type,
                self.localized_strings.get("longueur"),
                feature_attribute.length,
                self.localized_strings.get("precision"),
                feature_attribute.precision,
            )

        return out_attributes_str

    @lru_cache
    def format_size(self, in_size_in_octets: int = 0) -> str:
        """Format size in octets accordingly to the option.

        Args:
            in_size_in_octets (int): input size in octets

        Returns:
            str: formatted size in octets
        """
        if self.opt_size_prettify:
            return convert_octets(in_size_in_octets)
        else:
            return in_size_in_octets

    def store_error(
        self,
        metadataset: MetaDataset,
        worksheet: Worksheet,
        row_index: int,
    ):
        """Helper to store processing error in worksheet's row.

        Args:
            metadataset (MetaDataset): metadata with processing error
            worksheet (Worksheet): Excel workbook's sheet where to store
            row_index (int): worksheet's row index
        """
        err_mess = self.localized_strings.get(
            metadataset.processing_error_type, metadataset.processing_error_msg
        )
        worksheet[f"A{row_index}"] = metadataset.name
        worksheet[f"A{row_index}"].style = "Warning Text"
        worksheet[f"B{row_index}"] = self.format_as_hyperlink(
            target=metadataset.path_as_str,
            label=self.localized_strings.get("browse"),
        )
        worksheet[f"B{row_index}"].style = "Warning Text"
        worksheet[f"C{row_index}"] = err_mess
        worksheet[f"C{row_index}"].style = "Warning Text"
        # gdal info
        worksheet[f"Q{row_index}"] = (
            f"{metadataset.processing_error_type}: {metadataset.processing_error_msg}"
        )
        worksheet[f"Q{row_index}"].style = "Warning Text"
        logger.debug(
            f"Processing error detected on {metadataset.name} (in "
            f"{metadataset.path_as_str}) ({err_mess}) has been stored."
        )

    def get_sheet_and_incremented_row_index_from_type(
        self,
        metadataset: MetaDataset,
    ) -> tuple[Worksheet, int]:
        """Helper method to get the worksheet and row index corresponding to a type of
        metadataset.

        Args:
            metadataset: input metadataset

        Returns:
            worksheet, row index
        """
        if (
            isinstance(metadataset, MetaVectorDataset)
            and metadataset.dataset_type == "flat_vector"
        ):
            self.row_index_vector_files += 1
            return self.sheet_vector_files, self.row_index_vector_files
        elif isinstance(metadataset, MetaRasterDataset):
            self.row_index_raster_files += 1
            return self.sheet_raster_files, self.row_index_raster_files
        elif isinstance(metadataset, MetaDatabaseTable):
            self.row_index_server_geodatabases += 1
            return self.sheet_server_geodatabases, self.row_index_server_geodatabases
        elif isinstance(metadataset, MetaDatabaseFlat):
            self.row_index_flat_geodatabases += 1
            return self.sheet_flat_geodatabases, self.row_index_flat_geodatabases

    def serialize_metadaset(self, metadataset: MetaDataset) -> bool:
        """Router method to serialize metadataset depending on its type.

        Args:
            metadataset (MetaDataset): metadaset object to serialize

        Returns:
            bool: True if the object has been correctly serialized
        """
        worksheet, row_index = self.get_sheet_and_incremented_row_index_from_type(
            metadataset=metadataset
        )

        # -- Common --
        # in case of a source error
        if metadataset.processing_succeeded is False:
            self.store_error(
                metadataset=metadataset, worksheet=worksheet, row_index=row_index
            )

        # Dataset name
        worksheet[f"A{row_index}"] = metadataset.name

        # routing to custom serializing methods
        if (
            isinstance(metadataset, MetaVectorDataset)
            and metadataset.dataset_type == "flat_vector"
        ):
            return self.store_md_vector_files(
                metadataset=metadataset,
                worksheet=worksheet,
                row_index=row_index,
            )
        elif (
            isinstance(metadataset, MetaRasterDataset)
            and metadataset.dataset_type == "flat_raster"
        ):
            return self.store_md_raster_files(
                metadataset=metadataset,
                worksheet=worksheet,
                row_index=row_index,
            )
        elif (
            isinstance(metadataset, MetaDatabaseTable)
            and metadataset.dataset_type == "sgbd_postgis"
        ):
            return self.store_md_geodatabases_server(
                metadataset=metadataset,
                worksheet=worksheet,
                row_index=row_index,
            )
        elif isinstance(metadataset, MetaDatabaseFlat) and metadataset.dataset_type in (
            "flat_database",
            "flat_database_esri",
        ):
            return self.store_md_flat_geodatabases(
                metadataset=metadataset,
                worksheet=worksheet,
                row_index=row_index,
            )
        else:
            logger.error(
                f"No matching serializer for {metadataset.name} "
                f"(in {metadataset.path_as_str}): {metadataset.dataset_type}"
            )

    # ------------ Writing metadata ---------------------
    def store_md_vector_files(
        self,
        metadataset: MetaVectorDataset,
        worksheet: Worksheet | None = None,
        row_index: int | None = None,
    ):
        """Serialize a vector metadataset into an Excel worksheet's row.

        Args:
            metadataset (MetaDatabaseTable): metadataset to serialize
            worksheet (Workbook | None, optional): Excel workbook's sheet where to store. Defaults to None.
            row_index (int | None, optional): worksheet's row index. Defaults to None.
        """
        # if args not defined use class attributes
        if worksheet is None:
            worksheet = self.sheet_vector_files
        if row_index is None:
            row_index = self.row_index_vector_files

        # Path of parent folder formatted to be a hyperlink
        if self.opt_raw_path:
            worksheet[f"B{row_index}"] = metadataset.path_as_str
        else:
            worksheet[f"B{row_index}"] = self.format_as_hyperlink(
                target=metadataset.path.parent,
                label=self.localized_strings.get("browse"),
            )
            worksheet[f"B{row_index}"].style = "Hyperlink"

        # Name of parent folder with an exception if this is the format name
        worksheet[f"C{row_index}"] = metadataset.parent_folder_name

        # Fields count
        worksheet[f"D{row_index}"] = metadataset.count_feature_attributes
        # Objects count
        worksheet[f"E{row_index}"] = metadataset.features_objects_count
        # Geometry type
        worksheet[f"F{row_index}"] = metadataset.geometry_type

        # Name of srs
        worksheet[f"G{row_index}"] = metadataset.crs_name

        # Type of SRS
        worksheet[f"H{row_index}"] = self.localized_strings.get(
            metadataset.crs_type, str(metadataset.crs_type)
        )
        # SRS code
        worksheet[f"I{row_index}"] = (
            f"{metadataset.crs_registry}:{metadataset.crs_registry_code}"
        )
        # Spatial extent
        worksheet[f"J{row_index}"].style = "wrap"
        worksheet[f"J{row_index}"] = self.format_bbox(bbox=metadataset.bbox)

        # Creation date
        worksheet[f"K{row_index}"] = metadataset.storage_date_created
        # Last update date
        worksheet[f"L{row_index}"] = metadataset.storage_date_updated
        # Format of data
        worksheet[f"M{row_index}"] = metadataset.format_gdal_long_name
        # dependencies
        worksheet[f"N{row_index}"].style = "wrap"
        worksheet[f"N{row_index}"] = " |\n ".join(
            str(f.resolve()) for f in metadataset.files_dependencies
        )
        # total size
        worksheet[f"O{row_index}"] = self.format_size(
            in_size_in_octets=metadataset.storage_size
        )

        # Field informations
        worksheet[f"P{row_index}"] = self.format_feature_attributes(
            metadataset=metadataset
        )

    def store_md_raster_files(
        self,
        metadataset: MetaRasterDataset,
        worksheet: Worksheet | None = None,
        row_index: int | None = None,
    ):
        """Serialize a raster metadataset into an Excel worksheet's row.

        Args:
            metadataset (MetaDatabaseTable): metadataset to serialize
            worksheet (Workbook | None, optional): Excel workbook's sheet where to store. Defaults to None.
            row_index (int | None, optional): worksheet's row index. Defaults to None.
        """
        # if args not defined use class attributes
        if worksheet is None:
            worksheet = self.sheet_raster_files
        if row_index is None:
            row_index = self.row_index_raster_files

        # Path of parent folder formatted to be a hyperlink
        if self.opt_raw_path:
            worksheet[f"B{row_index}"] = metadataset.path_as_str
        else:
            worksheet[f"B{row_index}"] = self.format_as_hyperlink(
                target=metadataset.path.parent,
                label=self.localized_strings.get("browse"),
            )
            worksheet[f"B{row_index}"].style = "Hyperlink"

        # Name of parent folder with an exception if this is the format name
        worksheet[f"C{row_index}"] = metadataset.parent_folder_name

        # Image dimensions
        worksheet[f"D{row_index}"] = metadataset.rows_count
        worksheet[f"E{row_index}"] = metadataset.columns_count

        # Pixel dimensions
        worksheet[f"F{row_index}"] = metadataset.pixel_width
        worksheet[f"G{row_index}"] = metadataset.pixel_height

        # Image dimensions
        worksheet[f"H{row_index}"] = metadataset.origin_x
        worksheet[f"I{row_index}"] = metadataset.origin_y

        # Name of srs
        worksheet[f"J{row_index}"] = metadataset.crs_name

        # Type of SRS
        worksheet[f"K{row_index}"] = self.localized_strings.get(
            metadataset.crs_type, str(metadataset.crs_type)
        )
        # SRS code
        worksheet[f"L{row_index}"] = metadataset.crs_registry_code

        # Spatial extent
        worksheet[f"M{row_index}"].style = "wrap"
        worksheet[f"M{row_index}"] = self.format_bbox(bbox=metadataset.bbox)

        # Creation date
        worksheet[f"N{row_index}"] = metadataset.storage_date_created
        # Last update date
        worksheet[f"O{row_index}"] = metadataset.storage_date_updated

        # Number of bands
        worksheet[f"P{row_index}"] = metadataset.bands_count

        # Format of data
        worksheet[f"Q{row_index}"] = metadataset.format_gdal_long_name
        # Compression rate
        worksheet[f"R{row_index}"] = metadataset.compression_rate

        # Color referential
        worksheet[f"S{row_index}"] = metadataset.color_space

        # Dependencies
        worksheet[f"T{row_index}"].style = "wrap"
        worksheet[f"U{row_index}"] = " |\n ".join(
            f.resolve() for f in metadataset.files_dependencies
        )

        # total size of file and its dependencies
        worksheet[f"V{row_index}"] = self.format_size(
            in_size_in_octets=metadataset.storage_size
        )

    def store_md_flat_geodatabases(
        self,
        metadataset: MetaDatabaseFlat,
        worksheet: Worksheet | None = None,
        row_index: int | None = None,
    ):
        """Serialize a metadataset into an Excel worksheet's row.

        Args:
            metadataset (MetaDatabaseTable): metadataset to serialize
            worksheet (Workbook | None, optional): Excel workbook's sheet where to store. Defaults to None.
            row_index (int | None, optional): worksheet's row index. Defaults to None.
        """
        # if args not defined use class attributes
        if worksheet is None:
            worksheet = self.sheet_flat_geodatabases
        if row_index is None:
            row_index = self.row_index_flat_geodatabases

        # writing metadata
        if self.opt_raw_path:
            worksheet[f"B{row_index}"] = metadataset.path_as_str
        else:
            worksheet[f"B{row_index}"] = self.format_as_hyperlink(
                target=metadataset.path.parent,
                label=self.localized_strings.get("browse"),
            )
            worksheet[f"B{row_index}"].style = "Hyperlink"
        worksheet[f"C{row_index}"] = metadataset.parent_folder_name
        worksheet[f"D{row_index}"] = metadataset.storage_size
        worksheet[f"E{row_index}"] = metadataset.storage_date_created
        worksheet[f"F{row_index}"] = metadataset.storage_date_updated
        worksheet[f"G{row_index}"] = metadataset.format_gdal_long_name
        worksheet[f"H{row_index}"] = metadataset.count_layers
        worksheet[f"I{row_index}"] = metadataset.cumulated_count_feature_attributes
        worksheet[f"J{row_index}"] = metadataset.cumulated_count_feature_objects

        # parsing layers
        if not isinstance(metadataset.layers, Iterable):
            logger.warning(
                f"'{metadataset.name}' ({metadataset.path_as_str}) has no layers."
            )
            return

        for layer_metadataset in metadataset.layers:
            # increment line
            self.row_index_flat_geodatabases += 1
            row_index = self.row_index_flat_geodatabases

            # in case of a source error
            if metadataset.processing_succeeded is False:
                err_mess = self.localized_strings.get(
                    layer_metadataset.processing_error_type
                )
                logger.warning(
                    f"Problem detected on layer {layer_metadataset.name} "
                    f"(part of dataset '{metadataset.path_as_str}'). "
                    f"Error: {err_mess}"
                )
                worksheet[f"H{row_index}"] = layer_metadataset.name
                worksheet[f"H{row_index}"].style = "Warning Text"
                worksheet[f"I{row_index}"] = err_mess
                worksheet[f"I{row_index}"].style = "Warning Text"

            worksheet[f"H{row_index}"] = layer_metadataset.name
            worksheet[f"I{row_index}"] = layer_metadataset.count_feature_attributes
            worksheet[f"J{row_index}"] = layer_metadataset.features_objects_count
            worksheet[f"K{row_index}"] = layer_metadataset.geometry_type
            worksheet[f"L{row_index}"] = layer_metadataset.crs_name
            worksheet[f"M{row_index}"] = self.localized_strings.get(
                layer_metadataset.crs_type, str(layer_metadataset.crs_type)
            )

            # SRS code
            worksheet[f"N{row_index}"] = (
                f"{layer_metadataset.crs_registry}:{layer_metadataset.crs_registry_code}"
            )

            # Spatial extent
            worksheet[f"O{row_index}"].style = "wrap"
            worksheet[f"O{row_index}"] = self.format_bbox(bbox=layer_metadataset.bbox)

            # Field informations
            worksheet[f"P{row_index}"] = self.format_feature_attributes(
                metadataset=layer_metadataset
            )

    def store_md_mapdoc(
        self,
        metadataset: MetaVectorDataset,
        worksheet: Worksheet | None = None,
        row_index: int | None = None,
    ):
        """Serialize a metadataset into an Excel worksheet's row.

        Args:
            metadataset (MetaDatabaseTable): metadataset to serialize
            worksheet (Workbook | None, optional): Excel workbook's sheet where to store. Defaults to None.
            row_index (int | None, optional): worksheet's row index. Defaults to None.
        """
        # if args not defined use class attributes
        if worksheet is None:
            worksheet = self.sheet_map_workspaces
        if row_index is None:
            row_index = self.row_index_map_worskpaces

        # local variables
        champs = ""

        # Path of parent folder formatted to be a hyperlink
        if self.opt_raw_path:
            worksheet[f"B{row_index}"] = metadataset.path_as_str
        else:
            worksheet[f"B{row_index}"] = self.format_as_hyperlink(
                target=metadataset.path.parent,
                label=self.localized_strings.get("browse"),
            )
            worksheet[f"B{row_index}"].style = "Hyperlink"

        worksheet[f"C{row_index}"] = metadataset.parent_folder_name
        worksheet[f"E{row_index}"] = metadataset.get("creator_prod")
        worksheet[f"F{row_index}"] = metadataset.tags
        worksheet[f"G{row_index}"] = metadataset.subject
        worksheet[f"H{row_index}"] = metadataset.print_resolution_dpi
        worksheet[f"I{row_index}"] = metadataset.storage_size
        worksheet[f"J{row_index}"] = metadataset.storage_date_created
        worksheet[f"K{row_index}"] = metadataset.storage_date_updated
        worksheet[f"L{row_index}"] = metadataset.get("xOrigin")
        worksheet[f"M{row_index}"] = metadataset.get("yOrigin")
        worksheet[f"N{row_index}"] = self.localized_strings.get(
            metadataset.crs_name, ""
        )
        worksheet[f"O{row_index}"] = self.localized_strings.get(
            metadataset.crs_type, ""
        )
        worksheet[f"P{row_index}"] = self.localized_strings.get(
            metadataset.crs_registry_code, ""
        )
        worksheet[f"Q{row_index}"] = metadataset.format_gdal_long_name
        worksheet[f"R{row_index}"] = metadataset.count_feature_attributes
        worksheet[f"S{row_index}"] = metadataset.features_objects_count

        for layer_idx, layer_name in zip(
            metadataset.get("layers_idx"), metadataset.get("layers_names")
        ):
            # increment line
            self.row_index_map_worskpaces += 1
            champs = ""

            # get the layer informations
            try:
                mdoc_layer = metadataset.get(f"{layer_idx}_{layer_name}")
            except UnicodeDecodeError:
                mdoc_layer = metadataset.get(
                    "{}_{}".format(layer_idx, layer_name.encode("utf8", "replace"))
                )
            # in case of a source error
            if mdoc_layer.get("error"):
                err_mess = self.localized_strings.get(mdoc_layer.get("error"))
                logger.warning(
                    "Problem detected: {} in {}".format(
                        err_mess, mdoc_layer.get("title")
                    )
                )
                worksheet[f"Q{row_index}"] = mdoc_layer.get("title")
                worksheet[f"Q{row_index}"].style = "Warning Text"
                worksheet[f"R{row_index}"] = err_mess
                worksheet[f"R{row_index}"].style = "Warning Text"
                # loop must go on
                continue
            else:
                pass
            # layer info
            worksheet[f"Q{row_index}"] = mdoc_layer.get("title")
            worksheet[f"R{row_index}"] = mdoc_layer.get("num_fields")
            worksheet[f"S{row_index}"] = mdoc_layer.get("num_objs")

            # Field informations
            fields = mdoc_layer.get("fields")
            for chp in fields.keys():
                # field type
                if "Integer" in fields[chp][0]:
                    tipo = self.localized_strings.get("entier")
                elif fields[chp][0] == "Real":
                    tipo = self.localized_strings.get("reel")
                elif fields[chp][0] == "String":
                    tipo = self.localized_strings.get("string")
                elif fields[chp][0] == "Date":
                    tipo = self.localized_strings.get("date")
                else:
                    tipo = "unknown"
                    logger.warning(chp + " unknown type")

                # concatenation of field informations
                try:
                    champs = "{} {} ({}{}{}{}{}) ; ".format(
                        champs,
                        chp.decode("utf8", "replace"),
                        tipo,
                        self.localized_strings.get("longueur"),
                        fields[chp][1],
                        self.localized_strings.get("precision"),
                        fields[chp][2],
                    )
                except UnicodeDecodeError:
                    logger.warning(
                        "Field name with special letters: {}".format(
                            chp.decode("latin1")
                        )
                    )
                    # decode the fucking field name
                    champs = (
                        champs
                        + chp.decode("latin1")
                        + " ({}, Lg. = {}, Pr. = {}) ;".format(
                            tipo, fields[chp][1], fields[chp][2]
                        )
                    )
                    # then continue
                    continue

            # Once all fieds explored, write them
            self.sheet_flat_geodatabases[f"T{row_index}"] = champs

    def store_md_geodatabases_server(
        self,
        metadataset: MetaDatabaseTable,
        worksheet: Worksheet | None = None,
        row_index: int | None = None,
    ):
        """Serialize a metadataset into an Excel worksheet's row.

        Args:
            metadataset (MetaDatabaseTable): metadataset to serialize
            worksheet (Workbook | None, optional): Excel workbook's sheet where to store. Defaults to None.
            row_index (int | None, optional): worksheet's row index. Defaults to None.
        """
        # if args not defined use class attributes
        if worksheet is None:
            worksheet = self.sheet_server_geodatabases
        if row_index is None:
            row_index = self.row_index_server_geodatabases

        # connection string
        worksheet[f"B{row_index}"] = metadataset.path_as_str
        worksheet[f"B{row_index}"].style = "Hyperlink"
        # schema
        worksheet[f"C{row_index}"] = metadataset.schema_name

        # structure
        worksheet[f"D{row_index}"] = metadataset.count_feature_attributes
        worksheet[f"E{row_index}"] = metadataset.features_objects_count
        worksheet[f"F{row_index}"] = metadataset.geometry_type

        # SRS
        # Name of srs
        worksheet[f"G{row_index}"] = metadataset.crs_name
        # Type of SRS
        worksheet[f"H{row_index}"] = self.localized_strings.get(
            metadataset.crs_type, str(metadataset.crs_type)
        )
        # SRS code
        worksheet[f"I{row_index}"] = metadataset.crs_registry_code

        # Spatial extent
        worksheet[f"J{row_index}"].style = "wrap"
        worksheet[f"J{row_index}"] = self.format_bbox(bbox=metadataset.bbox)

        # type
        worksheet[f"K{row_index}"] = metadataset.format_gdal_long_name

        # Field informations
        worksheet[f"L{row_index}"] = self.format_feature_attributes(
            metadataset=metadataset
        )

#! python3  # noqa: E265

# ##############################################################################
# ########## Libraries #############
# ##################################

# Standard library
import logging
from os import path

# 3rd party library
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.utils import get_column_letter

# project
from dicogis.utils.formatters import convert_octets

# ##############################################################################
# ############ Globals ############
# #################################

# LOG
logger = logging.getLogger(__name__)

# ##############################################################################
# ########## Functions #############
# ##################################


def secure_encoding(layer_dict: dict, key_str: str) -> str:
    """Check if dictionary value is compatible with XML encoding.

    Args:
        layer_dict (dict): layer metadata dictionary
        key_str (str): key fo dictionary to check

    Returns:
        str: clean string
    """
    try:
        out_str = layer_dict.get(key_str, "").encode("utf-8", "strict")
        return out_str
    except UnicodeError as err:
        err_msg = "Encoding error spotted in '{}' for layer {}".format(
            key_str, layer_dict.get("name")
        )
        logger.warning(
            "{}. Layer: {}. Trace: {}".format(err_msg, layer_dict.get("name"), err)
        )
        return layer_dict.get(key_str, "").encode("utf8", "xmlcharrefreplace")


# ##############################################################################
# ########## Classes ###############
# ##################################


class MetadataToXlsx(Workbook):
    """Export into a XLSX worksheet."""

    li_cols_vector = [
        "nomfic",
        "path",
        "theme",
        "num_attrib",
        "num_objets",
        "geometrie",
        "srs",
        "srs_type",
        "codepsg",
        "emprise",
        "date_crea",
        "date_actu",
        "format",
        "li_depends",
        "tot_size",
        "li_chps",
        "gdal_warn",
    ]

    li_cols_raster = [
        "nomfic",
        "path",
        "theme",
        "size_Y",
        "size_X",
        "pixel_w",
        "pixel_h",
        "origin_x",
        "origin_y",
        "srs_type",
        "codepsg",
        "emprise",
        "date_crea",
        "date_actu",
        "num_bands",
        "format",
        "compression",
        "coloref",
        "li_depends",
        "tot_size",
        "gdal_warn",
    ]

    li_cols_filedb = [
        "nomfic",
        "path",
        "theme",
        "tot_size",
        "date_crea",
        "date_actu",
        "feats_class",
        "num_attrib",
        "num_objets",
        "geometrie",
        "srs",
        "srs_type",
        "codepsg",
        "emprise",
        "li_chps",
    ]

    li_cols_mapdocs = [
        "nomfic",
        "path",
        "theme",
        "custom_title",
        "creator_prod",
        "keywords",
        "subject",
        "res_image",
        "tot_size",
        "date_crea",
        "date_actu",
        "origin_x",
        "origin_y",
        "srs",
        "srs_type",
        "codepsg",
        "sub_layers",
        "num_attrib",
        "num_objets",
        "li_chps",
    ]

    li_cols_caodao = [
        "nomfic",
        "path",
        "theme",
        "tot_size",
        "date_crea",
        "date_actu",
        "sub_layers",
        "num_attrib",
        "num_objets",
        "geometrie",
        "srs",
        "srs_type",
        "codepsg",
        "emprise",
        "li_chps",
    ]

    li_cols_sgbd = [
        "nomfic",
        "conn_chain",
        "schema",
        "num_attrib",
        "num_objets",
        "geometrie",
        "srs",
        "srs_type",
        "codepsg",
        "emprise",
        "format",
        "li_chps",
        "gdal_err",
    ]

    def __init__(self, texts: dict, opt_size_prettify: bool = True):
        """Store metadata into Excel worksheets.

        Args:
            texts (dict, optional): dictionary of translated texts. Defaults to {}.
            opt_size_prettify (bool, optional): option to prettify size or not. Defaults to False.
        """
        super().__init__()
        self.txt = texts

        # options
        self.opt_size_prettify = opt_size_prettify

        # styles
        s_date = NamedStyle(name="date")
        s_date.number_format = "dd/mm/yyyy"
        s_wrap = NamedStyle(name="wrap")
        s_wrap.alignment = Alignment(wrap_text=True)
        self.add_named_style(s_date)
        self.add_named_style(s_wrap)

        # deleting the default worksheet
        ws = self.active
        self.remove_sheet(ws)

    # ------------ Setting workbook ---------------------

    def set_worksheets(
        self,
        has_vector=0,
        has_raster=0,
        has_filedb=0,
        has_mapdocs=0,
        has_cad=0,
        has_sgbd=0,
    ):
        """Add news sheets depending on present metadata types."""
        # SHEETS & HEADERS
        if has_vector and self.txt.get("sheet_vectors") not in self.sheetnames:
            self.ws_v = self.create_sheet(title=self.txt.get("sheet_vectors"))
            # headers
            self.ws_v.append([self.txt.get(i) for i in self.li_cols_vector])
            # styling
            for i in self.li_cols_vector:
                self.ws_v.cell(
                    row=1, column=self.li_cols_vector.index(i) + 1
                ).style = "Headline 2"

            # initialize line counter
            self.idx_v = 1
        else:
            pass

        if has_raster and self.txt.get("sheet_rasters") not in self.sheetnames:
            self.ws_r = self.create_sheet(title=self.txt.get("sheet_rasters"))
            # headers
            self.ws_r.append([self.txt.get(i) for i in self.li_cols_raster])
            # styling
            for i in self.li_cols_raster:
                self.ws_r.cell(
                    row=1, column=self.li_cols_raster.index(i) + 1
                ).style = "Headline 2"

            # initialize line counter
            self.idx_r = 1
        else:
            pass

        if has_filedb and self.txt.get("sheet_filedb") not in self.sheetnames:
            self.ws_fdb = self.create_sheet(title=self.txt.get("sheet_filedb"))
            # headers
            self.ws_fdb.append([self.txt.get(i) for i in self.li_cols_filedb])
            # styling
            for i in self.li_cols_filedb:
                self.ws_fdb.cell(
                    row=1, column=self.li_cols_filedb.index(i) + 1
                ).style = "Headline 2"

            # initialize line counter
            self.idx_f = 1
        else:
            pass

        if has_mapdocs and self.txt.get("sheet_maplans") not in self.sheetnames:
            self.ws_mdocs = self.create_sheet(title=self.txt.get("sheet_maplans"))
            # headers
            self.ws_mdocs.append([self.txt.get(i) for i in self.li_cols_mapdocs])
            # styling
            for i in self.li_cols_mapdocs:
                self.ws_mdocs.cell(
                    row=1, column=self.li_cols_mapdocs.index(i) + 1
                ).style = "Headline 2"

            # initialize line counter
            self.idx_m = 1
        else:
            pass

        if has_cad and self.txt.get("sheet_cdao") not in self.sheetnames:
            self.ws_cad = self.create_sheet(title=self.txt.get("sheet_cdao"))
            # headers
            self.ws_cad.append([self.txt.get(i) for i in self.li_cols_caodao])
            # styling
            for i in self.li_cols_caodao:
                self.ws_cad.cell(
                    row=1, column=self.li_cols_caodao.index(i) + 1
                ).style = "Headline 2"

            # initialize line counter
            self.idx_c = 1
        else:
            pass

        if has_sgbd and "PostGIS" not in self.sheetnames:
            self.ws_sgbd = self.create_sheet(title="PostGIS")
            # headers
            self.ws_sgbd.append([self.txt.get(i) for i in self.li_cols_sgbd])
            # styling
            for i in self.li_cols_sgbd:
                self.ws_sgbd.cell(
                    row=1, column=self.li_cols_sgbd.index(i) + 1
                ).style = "Headline 2"
            # initialize line counter
            self.idx_s = 1
        else:
            pass

        # end of method
        return

    def tunning_worksheets(self):
        """Clean up and tunning worksheet."""
        for sheet in self.worksheets:
            # Freezing panes
            c_freezed = sheet["B2"]
            sheet.freeze_panes = c_freezed

            # Print properties
            sheet.print_options.horizontalCentered = True
            sheet.print_options.verticalCentered = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

            # Others properties
            wsprops = sheet.sheet_properties
            wsprops.filterMode = True

            # enable filters
            sheet.auto_filter.ref = "A1:{}{}".format(
                get_column_letter(sheet.max_column), sheet.max_row
            )
            # columns width
            sheet.column_dimensions["A"].bestFit = True
            # sheet.column_dimensions['A'].auto_size = True
            # sheet.column_dimensions['B'].auto_size = True

            # dims = {}
            # for row in sheet.rows:
            #     for cell in row:
            #         if cell.value:
            #             val = cell.value
            #             dims[cell.column] = max((dims.get(cell.column, 0), len(val)))
            # for col, value in dims.items():
            #     sheet.column_dimensions[col].width = value
        # end of method
        return

    # ------------ Writing metadata ---------------------

    def store_md_vector(self, layer):
        """Store metadata about a vector dataset."""
        # increment line
        self.idx_v += 1
        # local variables
        champs = ""

        # in case of a source error
        if "error" in layer:
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.txt.get(layer.get("error"))
            logger.warning(
                "Problem detected: " "{} in {}".format(err_mess, layer.get("name"))
            )
            self.ws_v[f"A{self.idx_v}"] = layer.get("name")
            self.ws_v[f"A{self.idx_v}"].style = "Warning Text"
            link = r'=HYPERLINK("{}","{}")'.format(
                layer.get("folder"), self.txt.get("browse")
            )
            self.ws_v[f"B{self.idx_v}"] = link
            self.ws_v[f"B{self.idx_v}"].style = "Warning Text"
            self.ws_v[f"C{self.idx_v}"] = err_mess
            self.ws_v[f"C{self.idx_v}"].style = "Warning Text"
            # gdal info
            if "err_gdal" in layer:
                logger.warning(
                    "Problem detected by GDAL: "
                    "{} in {}".format(err_mess, layer.get("name"))
                )
                self.ws_v[f"Q{self.idx_v}"] = "{} : {}".format(
                    layer.get("err_gdal")[0], layer.get("err_gdal")[1]
                )
                self.ws_v[f"Q{self.idx_v}"].style = "Warning Text"
            else:
                pass
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_v[f"A{self.idx_v}"] = secure_encoding(layer, "name")

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{}","{}")'.format(
            layer.get("folder"), self.txt.get("browse")
        )
        self.ws_v[f"B{self.idx_v}"] = link
        self.ws_v[f"B{self.idx_v}"].style = "Hyperlink"

        # Name of parent folder with an exception if this is the format name
        self.ws_v[f"C{self.idx_v}"] = path.basename(layer.get("folder"))

        # Fields count
        self.ws_v[f"D{self.idx_v}"] = layer.get("num_fields", "")
        # Objects count
        self.ws_v[f"E{self.idx_v}"] = layer.get("num_obj", "")
        # Geometry type
        self.ws_v[f"F{self.idx_v}"] = layer.get("type_geom", "")

        # Name of srs
        self.ws_v[f"G{self.idx_v}"] = secure_encoding(layer, "srs")

        # Type of SRS
        self.ws_v[f"H{self.idx_v}"] = layer.get("srs_type", "")
        # EPSG code
        self.ws_v[f"I{self.idx_v}"] = layer.get("epsg", "")
        # Spatial extent
        emprise = "Xmin : {} - Xmax : {} | \nYmin : {} - Ymax : {}".format(
            layer.get("xmin"),
            layer.get("xmax"),
            layer.get("ymin"),
            layer.get("ymax"),
        )
        self.ws_v[f"J{self.idx_v}"].style = "wrap"
        self.ws_v[f"J{self.idx_v}"] = emprise

        # Creation date
        self.ws_v[f"K{self.idx_v}"] = layer.get("date_crea")
        # Last update date
        self.ws_v[f"L{self.idx_v}"] = layer.get("date_actu")
        # Format of data
        self.ws_v[f"M{self.idx_v}"] = layer.get("format")
        # dependencies
        self.ws_v[f"N{self.idx_v}"].style = "wrap"
        self.ws_v[f"N{self.idx_v}"] = " |\n ".join(layer.get("dependencies", []))
        # total size
        if self.opt_size_prettify:
            self.ws_v[f"O{self.idx_v}"] = convert_octets(layer.get("total_size", 0))
        else:
            self.ws_v[f"O{self.idx_v}"] = layer.get("total_size", 0)

        # Field informations
        fields = layer.get("fields")
        for chp in fields.keys():
            # field type
            if "Integer" in fields[chp][0]:
                tipo = self.txt.get("entier")
            elif fields[chp][0] == "Real":
                tipo = self.txt.get("reel")
            elif fields[chp][0] == "String":
                tipo = self.txt.get("string")
            elif fields[chp][0] == "Date":
                tipo = self.txt.get("date")
            else:
                tipo = "unknown"
                logger.warning(chp + " unknown type")

            # concatenation of field informations
            try:
                champs = "{} {} ({}{}{}{}{}) ; ".format(
                    champs,
                    chp.encode("utf8", "replace"),
                    tipo,
                    self.txt.get("longueur"),
                    fields[chp][1],
                    self.txt.get("precision"),
                    fields[chp][2],
                )
            except UnicodeDecodeError:
                logger.warning(f"Field name with special letters: {chp}")
                continue

        # Once all fieds explored, write them
        self.ws_v[f"P{self.idx_v}"] = champs

        # end of method
        return

    def store_md_raster(self, layer, bands):
        """Store metadata about a raster dataset."""
        # increment line
        self.idx_r += 1

        # in case of a source error
        if "error" in layer:
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.txt.get(layer.get("error"))
            logger.warning(
                "Problem detected: " "{} in {}".format(err_mess, layer.get("name"))
            )
            self.ws_r[f"A{self.idx_r}"] = layer.get("name")
            link = r'=HYPERLINK("{}","{}")'.format(
                layer.get("folder"), self.txt.get("browse")
            )
            self.ws_r[f"B{self.idx_r}"] = link
            self.ws_r[f"B{self.idx_r}"].style = "Warning Text"
            self.ws_r[f"C{self.idx_r}"] = err_mess
            self.ws_r[f"C{self.idx_r}"].style = "Warning Text"
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_r[f"A{self.idx_r}"] = layer.get("name")

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{}","{}")'.format(
            layer.get("folder"), self.txt.get("browse")
        )
        self.ws_r[f"B{self.idx_r}"] = link
        self.ws_r[f"B{self.idx_r}"].style = "Hyperlink"

        # Name of parent folder with an exception if this is the format name
        self.ws_r[f"C{self.idx_r}"] = path.basename(layer.get("folder"))

        # Image dimensions
        self.ws_r[f"D{self.idx_r}"] = layer.get("num_rows")
        self.ws_r[f"E{self.idx_r}"] = layer.get("num_cols")

        # Pixel dimensions
        self.ws_r[f"F{self.idx_r}"] = layer.get("pixelWidth")
        self.ws_r[f"G{self.idx_r}"] = layer.get("pixelHeight")

        # Image dimensions
        self.ws_r[f"H{self.idx_r}"] = layer.get("xOrigin")
        self.ws_r[f"I{self.idx_r}"] = layer.get("yOrigin")

        # Type of SRS
        self.ws_r[f"J{self.idx_r}"] = layer.get("srs_type")
        # EPSG code
        self.ws_r[f"K{self.idx_r}"] = layer.get("epsg")

        # Creation date
        self.ws_r[f"M{self.idx_r}"] = layer.get("date_crea")
        # Last update date
        self.ws_r[f"N{self.idx_r}"] = layer.get("date_actu")

        # Number of bands
        self.ws_r[f"O{self.idx_r}"] = layer.get("num_bands")

        # Format of data
        self.ws_r[f"P{self.idx_r}"] = "{} {}".format(
            layer.get("format"), layer.get("format_version")
        )
        # Compression rate
        self.ws_r[f"Q{self.idx_r}"] = layer.get("compr_rate")

        # Color referential
        self.ws_r[f"R{self.idx_r}"] = layer.get("color_ref")

        # Dependencies
        self.ws_r[f"S{self.idx_v}"].style = "wrap"
        self.ws_r[f"S{self.idx_v}"] = " |\n ".join(layer.get("dependencies"))

        # total size of file and its dependencies
        if self.opt_size_prettify:
            try:
                self.ws_v[f"O{self.idx_r}"] = convert_octets(layer.get("total_size", 0))
            except TypeError as exc:
                logger.error(
                    f"Failed to prettify size of {layer.get('name')}. Trace: {exc}"
                )

        else:
            self.ws_v[f"O{self.idx_r}"] = layer.get("total_size", 0)

        # in case of a source error
        if layer.get("err_gdal", [0])[0] != 0:
            logger.warning(
                "Problem detected by GDAL: "
                "{} in {}".format(layer.get("err_gdal"), layer.get("name"))
            )
            self.ws_r[f"U{self.idx_r}"] = "{} : {}".format(
                layer.get("err_gdal")[0], layer.get("err_gdal")[1]
            )
            self.ws_r[f"U{self.idx_r}"].style = "Warning Text"
        else:
            pass

        # end of method
        return

    def store_md_fdb(self, filedb):
        """Storing metadata about a file database."""
        # increment line
        self.idx_f += 1

        # in case of a source error
        if "error" in filedb:
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.txt.get(filedb.get("error"))
            logger.warning(
                "Problem detected: " "{} in {}".format(err_mess, filedb.get("name"))
            )
            self.ws_fdb[f"A{self.idx_f}"] = filedb.get("name")
            link = r'=HYPERLINK("{}","{}")'.format(
                filedb.get("folder"), self.txt.get("browse")
            )
            self.ws_fdb[f"B{self.idx_f}"] = link
            self.ws_fdb[f"B{self.idx_f}"].style = "Warning Text"
            self.ws_fdb[f"C{self.idx_f}"] = err_mess
            self.ws_fdb[f"C{self.idx_f}"].style = "Warning Text"
            # gdal info
            if "err_gdal" in filedb:
                logger.warning(
                    "Problem detected by GDAL: "
                    "{} in {}".format(err_mess, filedb.get("name"))
                )
                self.ws_fdb[f"Q{self.idx_v}"] = "{} : {}".format(
                    filedb.get("err_gdal")[0], filedb.get("err_gdal")[1]
                )
                self.ws_fdb[f"Q{self.idx_v}"].style = "Warning Text"
            else:
                pass
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_fdb[f"A{self.idx_f}"] = filedb.get("name")

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{}","{}")'.format(
            filedb.get("folder"), self.txt.get("browse")
        )
        self.ws_fdb[f"B{self.idx_f}"] = link
        self.ws_fdb[f"B{self.idx_f}"].style = "Hyperlink"

        self.ws_fdb[f"C{self.idx_f}"] = path.basename(filedb.get("folder"))
        self.ws_fdb[f"D{self.idx_f}"] = filedb.get("total_size")
        self.ws_fdb[f"E{self.idx_f}"] = filedb.get("date_crea")
        self.ws_fdb[f"F{self.idx_f}"] = filedb.get("date_actu")
        self.ws_fdb[f"G{self.idx_f}"] = filedb.get("layers_count")
        self.ws_fdb[f"H{self.idx_f}"] = filedb.get("total_fields")
        self.ws_fdb[f"I{self.idx_f}"] = filedb.get("total_objs")

        # parsing layers
        for layer_idx, layer_name in zip(
            filedb.get("layers_idx"), filedb.get("layers_names")
        ):
            # increment line
            self.idx_f += 1
            champs = ""
            # get the layer informations
            try:
                gdb_layer = filedb.get(f"{layer_idx}_{layer_name}")
            except UnicodeError as err:
                logger.error(f"Encoding error. Trace: {err}")
                continue
            # in case of a source error
            if gdb_layer.get("error"):
                err_mess = self.txt.get(gdb_layer.get("error"))
                logger.warning(
                    "Problem detected: {} in {}".format(
                        err_mess, gdb_layer.get("title")
                    )
                )
                self.ws_fdb[f"G{self.idx_f}"] = gdb_layer.get("title")
                self.ws_fdb[f"G{self.idx_f}"].style = "Warning Text"
                self.ws_fdb[f"H{self.idx_f}"] = err_mess
                self.ws_fdb[f"H{self.idx_f}"].style = "Warning Text"
                # Interruption of function
                continue
            else:
                pass

            self.ws_fdb[f"G{self.idx_f}"] = gdb_layer.get("title")
            self.ws_fdb[f"H{self.idx_f}"] = gdb_layer.get("num_fields")
            self.ws_fdb[f"I{self.idx_f}"] = gdb_layer.get("num_obj")
            self.ws_fdb[f"J{self.idx_f}"] = gdb_layer.get("type_geom")
            self.ws_fdb[f"K{self.idx_f}"] = secure_encoding(gdb_layer, "srs")
            self.ws_fdb[f"L{self.idx_f}"] = gdb_layer.get("srs_type")
            self.ws_fdb[f"M{self.idx_f}"] = gdb_layer.get("epsg")

            # Spatial extent
            emprise = "Xmin : {} - Xmax : {} | \nYmin : {} - Ymax : {}".format(
                gdb_layer.get("xmin"),
                gdb_layer.get("xmax"),
                gdb_layer.get("ymin"),
                gdb_layer.get("ymax"),
            )
            self.ws_fdb[f"N{self.idx_f}"].style = "wrap"
            self.ws_fdb[f"N{self.idx_f}"] = emprise

            # Field informations
            fields = gdb_layer.get("fields")
            for chp in fields.keys():
                # field type
                if "Integer" in fields[chp][0]:
                    tipo = self.txt.get("entier")
                elif fields[chp][0] == "Real":
                    tipo = self.txt.get("reel")
                elif fields[chp][0] == "String":
                    tipo = self.txt.get("string")
                elif fields[chp][0] == "Date":
                    tipo = self.txt.get("date")
                else:
                    tipo = "unknown"
                    logger.warning(chp + " unknown type")

                # concatenation of field informations
                try:
                    champs = "{} {} ({}{}{}{}{}) ; ".format(
                        champs,
                        chp.encode("utf8", "replace"),
                        tipo,
                        self.txt.get("longueur"),
                        fields[chp][1],
                        self.txt.get("precision"),
                        fields[chp][2],
                    )
                except UnicodeDecodeError:
                    logger.warning(f"Field name with special letters: {chp}")
                    continue

            # Once all fieds explored, write them
            self.ws_fdb[f"O{self.idx_f}"] = champs

        # end of method
        return

    def store_md_mapdoc(self, mapdoc):
        """To store mapdocs information from DicoGIS."""
        # increment line
        self.idx_m += 1

        # local variables
        champs = ""

        # in case of a source error
        if "error" in mapdoc:
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.txt.get(mapdoc.get("error"))
            logger.warning(
                "Problem detected: " "{} in {}".format(err_mess, mapdoc.get("name"))
            )
            self.ws_mdocs[f"A{self.idx_m}"] = mapdoc.get("name")
            self.ws_mdocs[f"A{self.idx_m}"].style = "Warning Text"
            link = r'=HYPERLINK("{}","{}")'.format(
                mapdoc.get("folder"), self.txt.get("browse")
            )
            self.ws_mdocs[f"B{self.idx_m}"] = link
            self.ws_mdocs[f"B{self.idx_m}"].style = "Warning Text"
            self.ws_mdocs[f"C{self.idx_m}"] = err_mess
            self.ws_mdocs[f"C{self.idx_m}"].style = "Warning Text"
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_mdocs[f"A{self.idx_m}"] = mapdoc.get("name")

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{}","{}")'.format(
            mapdoc.get("folder"), self.txt.get("browse")
        )
        self.ws_mdocs[f"B{self.idx_m}"] = link
        self.ws_mdocs[f"B{self.idx_m}"].style = "Hyperlink"
        self.ws_mdocs[f"C{self.idx_m}"] = path.dirname(mapdoc.get("folder"))
        self.ws_mdocs[f"D{self.idx_m}"] = mapdoc.get("title")
        self.ws_mdocs[f"E{self.idx_m}"] = mapdoc.get("creator_prod")
        self.ws_mdocs[f"F{self.idx_m}"] = mapdoc.get("keywords")
        self.ws_mdocs[f"G{self.idx_m}"] = mapdoc.get("subject")
        self.ws_mdocs[f"H{self.idx_m}"] = mapdoc.get("dpi")
        self.ws_mdocs[f"I{self.idx_m}"] = mapdoc.get("total_size")
        self.ws_mdocs[f"J{self.idx_m}"] = mapdoc.get("date_crea")
        self.ws_mdocs[f"K{self.idx_m}"] = mapdoc.get("date_actu")
        self.ws_mdocs[f"L{self.idx_m}"] = mapdoc.get("xOrigin")
        self.ws_mdocs[f"M{self.idx_m}"] = mapdoc.get("yOrigin")
        self.ws_mdocs[f"N{self.idx_m}"] = secure_encoding(mapdoc, "srs")
        self.ws_mdocs[f"O{self.idx_m}"] = mapdoc.get("srs_type")
        self.ws_mdocs[f"P{self.idx_m}"] = mapdoc.get("epsg")
        self.ws_mdocs[f"Q{self.idx_m}"] = mapdoc.get("layers_count")
        self.ws_mdocs[f"R{self.idx_m}"] = mapdoc.get("total_fields")
        self.ws_mdocs[f"S{self.idx_m}"] = mapdoc.get("total_objs")

        for layer_idx, layer_name in zip(
            mapdoc.get("layers_idx"), mapdoc.get("layers_names")
        ):
            # increment line
            self.idx_m += 1
            champs = ""

            # get the layer informations
            try:
                mdoc_layer = mapdoc.get(f"{layer_idx}_{layer_name}")
            except UnicodeDecodeError:
                mdoc_layer = mapdoc.get(
                    "{}_{}".format(layer_idx, layer_name.encode("utf8", "replace"))
                )
            # in case of a source error
            if mdoc_layer.get("error"):
                err_mess = self.txt.get(mdoc_layer.get("error"))
                logger.warning(
                    "Problem detected: {} in {}".format(
                        err_mess, mdoc_layer.get("title")
                    )
                )
                self.ws_mdocs[f"Q{self.idx_f}"] = mdoc_layer.get("title")
                self.ws_mdocs[f"Q{self.idx_f}"].style = "Warning Text"
                self.ws_mdocs[f"R{self.idx_f}"] = err_mess
                self.ws_mdocs[f"R{self.idx_f}"].style = "Warning Text"
                # loop must go on
                continue
            else:
                pass
            # layer info
            self.ws_mdocs[f"Q{self.idx_m}"] = mdoc_layer.get("title")
            self.ws_mdocs[f"R{self.idx_m}"] = mdoc_layer.get("num_fields")
            self.ws_mdocs[f"S{self.idx_m}"] = mdoc_layer.get("num_objs")

            # Field informations
            fields = mdoc_layer.get("fields")
            for chp in fields.keys():
                # field type
                if "Integer" in fields[chp][0]:
                    tipo = self.txt.get("entier")
                elif fields[chp][0] == "Real":
                    tipo = self.txt.get("reel")
                elif fields[chp][0] == "String":
                    tipo = self.txt.get("string")
                elif fields[chp][0] == "Date":
                    tipo = self.txt.get("date")
                else:
                    tipo = "unknown"
                    logger.warning(chp + " unknown type")

                # concatenation of field informations
                try:
                    champs = "{} {} ({}{}{}{}{}) ; ".format(
                        champs,
                        chp.decode("utf8", "replace"),
                        tipo,
                        self.txt.get("longueur"),
                        fields[chp][1],
                        self.txt.get("precision"),
                        fields[chp][2],
                    )
                except UnicodeDecodeError:
                    logger.warning(
                        "Field name with special letters: {}".format(
                            chp.decode("latin1")
                        )
                    )
                    # decode the fucking field name
                    champs = (
                        champs
                        + chp.decode("latin1")
                        + " ({}, Lg. = {}, Pr. = {}) ;".format(
                            tipo, fields[chp][1], fields[chp][2]
                        )
                    )
                    # then continue
                    continue

            # Once all fieds explored, write them
            self.ws_fdb[f"T{self.idx_f}"] = champs

        # end of method
        return

    def store_md_cad(self, cad):
        """Store metadata about CAD dataset."""
        # increment line
        self.idx_c += 1

        # local variables
        champs = ""

        # in case of a source error
        if "error" in cad:
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.txt.get(cad.get("error"))
            logger.warning(
                "Problem detected: {} in {}".format(err_mess, cad.get("name"))
            )
            self.ws_cad[f"A{self.idx_c}"] = cad.get("name")
            self.ws_cad[f"A{self.idx_c}"].style = "Warning Text"
            link = r'=HYPERLINK("{}","{}")'.format(
                cad.get("folder"), self.txt.get("browse")
            )
            self.ws_cad[f"B{self.idx_c}"] = link
            self.ws_cad[f"B{self.idx_c}"].style = "Warning Text"
            self.ws_cad[f"C{self.idx_c}"] = err_mess
            self.ws_cad[f"C{self.idx_c}"].style = "Warning Text"
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_cad[f"A{self.idx_c}"] = cad.get("name")

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{}","{}")'.format(
            cad.get("folder"), self.txt.get("browse")
        )
        self.ws_cad[f"B{self.idx_c}"] = link
        self.ws_cad[f"B{self.idx_c}"].style = "Hyperlink"

        # Name of parent folder with an exception if this is the format name
        self.ws_cad[f"C{self.idx_c}"] = path.basename(cad.get("folder"))
        # total size
        self.ws_cad[f"D{self.idx_c}"] = cad.get("total_size")
        # Creation date
        self.ws_cad[f"E{self.idx_c}"] = cad.get("date_crea")
        # Last update date
        self.ws_cad[f"F{self.idx_c}"] = cad.get("date_actu")
        self.ws_cad[f"G{self.idx_c}"] = cad.get("layers_count")
        self.ws_cad[f"H{self.idx_c}"] = cad.get("total_fields")
        self.ws_cad[f"I{self.idx_c}"] = cad.get("total_objs")

        # parsing layers
        for layer_idx, layer_name in zip(
            cad.get("layers_idx"), cad.get("layers_names")
        ):
            # increment line
            self.idx_c += 1
            champs = ""
            # get the layer informations
            try:
                layer = cad.get(f"{layer_idx}_{layer_name}")
            except UnicodeDecodeError:
                layer = cad.get("{}_{}".format(layer_idx, layer_name.decode("latin1")))
            # in case of a source error
            if layer.get("error"):
                err_mess = self.txt.get(layer.get("error"))
                logger.warning(
                    "Problem detected: " "{} in {}".format(err_mess, layer.get("title"))
                )
                self.ws_cad[f"G{self.idx_c}"] = layer.get("title")
                self.ws_cad[f"G{self.idx_c}"].style = "Warning Text"
                self.ws_cad[f"H{self.idx_c}"] = err_mess
                self.ws_cad[f"H{self.idx_c}"].style = "Warning Text"
                # Interruption of function
                continue
            else:
                pass

            self.ws_cad[f"G{self.idx_c}"] = layer.get("title")
            self.ws_cad[f"H{self.idx_c}"] = layer.get("num_fields")
            self.ws_cad[f"I{self.idx_c}"] = layer.get("num_obj")
            self.ws_cad[f"J{self.idx_c}"] = layer.get("type_geom")
            self.ws_cad[f"K{self.idx_c}"] = layer.get("srs")
            self.ws_cad[f"L{self.idx_c}"] = layer.get("srs_type")
            self.ws_cad[f"M{self.idx_c}"] = layer.get("epsg")

            # Spatial extent
            emprise = "Xmin : {} - Xmax : {} | \nYmin : {} - Ymax : {}".format(
                layer.get("xmin"),
                layer.get("xmax"),
                layer.get("ymin"),
                layer.get("ymax"),
            )
            self.ws_cad[f"N{self.idx_c}"].style = "wrap"
            self.ws_cad[f"N{self.idx_c}"] = emprise

            # Field informations
            fields = layer.get("fields")
            for chp in fields.keys():
                # field type
                if "Integer" in fields[chp][0]:
                    tipo = self.txt.get("entier")
                elif fields[chp][0] == "Real":
                    tipo = self.txt.get("reel")
                elif fields[chp][0] == "String":
                    tipo = self.txt.get("string")
                elif fields[chp][0] == "Date":
                    tipo = self.txt.get("date")
                else:
                    tipo = "unknown"
                    logger.warning(chp + " unknown type")

                # concatenation of field informations
                try:
                    champs = "{} {} ({}{}{}{}{}) ; ".format(
                        champs,
                        chp,
                        tipo,
                        self.txt.get("longueur"),
                        fields[chp][1],
                        self.txt.get("precision"),
                        fields[chp][2],
                    )
                except UnicodeDecodeError:
                    logger.warning(
                        "Field name with special letters: {}".format(
                            chp.decode("latin1")
                        )
                    )
                    # decode the fucking field name
                    champs = (
                        champs
                        + chp.decode("latin1")
                        + " ({}, Lg. = {}, Pr. = {}) ;".format(
                            tipo, fields[chp][1], fields[chp][2]
                        )
                    )
                    # then continue
                    continue

            # Once all fieds explored, write them
            self.ws_cad[f"O{self.idx_c}"] = champs

        # End of method
        return

    def store_md_sgdb(self, layer):
        """Storing metadata about a file database."""
        # increment line
        self.idx_s += 1
        # local variable
        champs = ""

        # layer name
        self.ws_sgbd[f"A{self.idx_s}"] = layer.get("name")

        # connection string
        self.ws_sgbd[f"B{self.idx_s}"] = "{}@{}:{}-{}".format(
            layer.get("user"),
            layer.get("sgbd_host"),
            layer.get("sgbd_port"),
            layer.get("db_name"),
        )
        self.ws_sgbd[f"B{self.idx_s}"].style = "Hyperlink"
        # schema
        self.ws_sgbd[f"C{self.idx_s}"] = layer.get("folder")

        # in case of a source error
        if "error" in layer:
            self.ws_sgbd[f"D{self.idx_s}"] = layer.get("error")
            self.ws_sgbd[f"A{self.idx_s}"].style = "Warning Text"
            self.ws_sgbd[f"B{self.idx_s}"].style = "Warning Text"
            self.ws_sgbd[f"C{self.idx_s}"].style = "Warning Text"
            self.ws_sgbd[f"D{self.idx_s}"].style = "Warning Text"
            # gdal info
            if "err_gdal" in layer:
                self.ws_v[f"M{self.idx_v}"] = "{} : {}".format(
                    layer.get("err_gdal")[0], layer.get("err_gdal")[1]
                )
                self.ws_v[f"M{self.idx_v}"].style = "Warning Text"
            else:
                pass
            # interruption of function
            return False
        else:
            pass

        # structure
        self.ws_sgbd[f"D{self.idx_s}"] = layer.get("num_fields")
        self.ws_sgbd[f"E{self.idx_s}"] = layer.get("num_obj")
        self.ws_sgbd[f"F{self.idx_s}"] = layer.get("type_geom")

        # SRS
        self.ws_sgbd[f"G{self.idx_s}"] = layer.get("srs")
        self.ws_sgbd[f"H{self.idx_s}"] = layer.get("srs_type")
        self.ws_sgbd[f"I{self.idx_s}"] = layer.get("epsg")

        # Spatial extent
        emprise = "Xmin : {} - Xmax : {} | \nYmin : {} - Ymax : {}".format(
            layer.get("xmin"),
            layer.get("xmax"),
            layer.get("ymin"),
            layer.get("ymax"),
        )
        self.ws_sgbd[f"J{self.idx_s}"].style = "wrap"
        self.ws_sgbd[f"J{self.idx_s}"] = emprise

        # type
        self.ws_sgbd[f"K{self.idx_s}"] = layer.get("format")

        # Field informations
        fields = layer.get("fields")
        for chp in fields.keys():
            # field type
            if "Integer" in fields[chp][0]:
                tipo = self.txt.get("entier")
            elif fields[chp][0] == "Real":
                tipo = self.txt.get("reel")
            elif fields[chp][0] == "String":
                tipo = self.txt.get("string")
            elif fields[chp][0] == "Date":
                tipo = self.txt.get("date")
            else:
                tipo = "unknown"
                logger.warning(chp + " unknown type")

            # concatenation of field informations
            try:
                champs = "{} {} ({}{}{}{}{}) ; ".format(
                    champs,
                    chp,
                    tipo,
                    self.txt.get("longueur"),
                    fields[chp][1],
                    self.txt.get("precision"),
                    fields[chp][2],
                )
            except UnicodeDecodeError:
                logger.warning(
                    "Field name with special letters: {}".format(chp.decode("latin1"))
                )
                # decode the fucking field name
                champs = (
                    champs
                    + chp.decode("latin1")
                    + " ({}, Lg. = {}, Pr. = {}) ;".format(
                        tipo, fields[chp][1], fields[chp][2]
                    )
                )
                # then continue
                continue

        # Once all fieds explored, write them
        self.ws_sgbd[f"L{self.idx_s}"] = champs

        # end of method
        return


# ############################################################################
# ##### Stand alone program ########
# ##################################
if __name__ == "__main__":
    """Standalone execution and tests"""
    pass
